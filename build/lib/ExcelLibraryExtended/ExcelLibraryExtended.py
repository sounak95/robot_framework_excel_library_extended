from openpyxl import load_workbook
from ExcelLibrary import ExcelLibrary
import time


class ExcelLibraryExtended(ExcelLibrary):
    """
        This test library provides keywords to allow opening, reading, writing, and saving Excel files from Unified Test Framework.
    """

    def edit_data_xlsx_file(self, file_path, sheetname, coloumnheader, changedata, rownumber=None):

        """| Usage |
                It updates the data in a xlsx file.

                | Arguments |

                'file_path' = xlsx file location.

                'sheetname' = SheetName of the xlsx file.

                'coloumnheader' = ColoumnHeader name.

                'changedata' = data to be updated in the xlsx file.

                'rownumber' [Optional]= If provided as an integer value, it will update the data into the corrosponding rownumber. By default it selects the first row of the corrosponding columnheader.

                Example:

                |***TestCases*** |

                1. To Update Data Into The First Row of The Corrosponding Columnheader :
                |Edit Data Xlsx File | file_path=C:/example.xlsx | sheetname=Sheet1 | coloumnheader=Sample | changedata=hello world  |

                2. To Update Data Into The Fourth Row of The Corrosponding Columnheader :
                |Edit Data Xlsx File | file_path=C:/example.xlsx | sheetname=Sheet1 | coloumnheader=Sample | changedata=hello world  |  rownumber= 4

               """
        wb = load_workbook(file_path)
        sheet = wb[sheetname]     # get Sheet
        r = sheet.max_row  # No of written Rows in sheet
        c = sheet.max_column  # No of written Columns in sheet
        for i in range(1, r + 1):  # Reading each cell in excel
            for j in range(1, c + 1):
                if str(sheet.cell(i, j).value).lower() == coloumnheader.lower():
                    if not rownumber:
                        sheet.cell(i + 1, j).value = changedata
                    else:
                        try:
                            rownumber = int(rownumber)
                            sheet.cell(rownumber, j).value = changedata
                        except:
                            raise AssertionError("Please Provide a valid 'rownumber' as an integer. Given rownumber is: {}".format(rownumber))
        time.sleep(0.3)
        try:
            wb.save(filename=file_path)
        except:
                raise AssertionError("Please close {} file manually to update data".format(file_path))
        wb.close()